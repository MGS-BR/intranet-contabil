import os
import re
from datetime import datetime
from pypdf import PdfReader
from openpyxl import load_workbook, Workbook

REGRAS_CONTAS_DESPESAS = [
    ("INSS SOBRE", "77"),
    ("INSS RETIDO", "77"),

    ("TECNOLOGIA E INOVACAO", "136"),
    ("TECNOLOGIA E INOVAÇÃO", "136"),

    ("OUTROS SERVICOS", "72"),
    ("OUTROS SERVIÇOS", "72"),

    ("SERVICOS DE ELETRICISTA", "128"),
    ("SERVIÇOS DE ELETRICISTA", "128"),

    ("AQUISICAO DE VENTILADORES", "251"),
    ("AQUISIÇÃO DE VENTILADORES", "251"),
    ("VENTILADORES", "251"),
    ("VENTILADOR", "251"),

    ("EQUIPAMENTOS PARA REFRIGERACAO", "251"),
    ("EQUIPAMENTOS PARA REFRIGERAÇÃO", "251"),
    ("CONDICIONAMENTO", "251"),
    ("CIRCULACAO DE AR", "251"),
    ("CIRCULAÇÃO DE AR", "251"),

    ("MANUTENCAO PARA EQUIPAMENTOS DE INFORMATICA", "71"),
    ("MANUTENÇÃO PARA EQUIPAMENTOS DE INFORMÁTICA", "71"),
    ("SERVICO DE MANUTENCAO EM EQUIPTO DE INFORMATICA", "71"),
    ("SERVIÇO DE MANUTENÇÃO EM EQUIPTO DE INFORMÁTICA", "71"),
    ("IMPRESSORA", "71"),

    ("SERVICOS GERAIS INTERNOS", "180"),
    ("SERVIÇOS GERAIS INTERNOS", "180"),
    ("SERVICOS GERAIS", "180"),
    ("SERVIÇOS GERAIS", "180"),

    ("SERVICOS DE DESINSETIZACAO", "128"),
    ("SERVIÇOS DE DESINSETIZAÇÃO", "128"),
    ("DESINSETIZACAO", "128"),
    ("DESINSETIZAÇÃO", "128"),

    ("SERVICOS DE IMUNIZACAO", "128"),
    ("SERVIÇOS DE IMUNIZAÇÃO", "128"),
    ("IMUNIZACAO", "128"),
    ("IMUNIZAÇÃO", "128"),
    ("HIGIENIZACAO", "128"),
    ("HIGIENIZAÇÃO", "128"),

    ("SERVICOS DE JARDINAGEM", "128"),
    ("SERVIÇOS DE JARDINAGEM", "128"),
    ("JARDINAGEM", "128"),

    ("SERVICOS DE ENCANADOR", "128"),
    ("SERVIÇOS DE ENCANADOR", "128"),
    ("ENCANADOR", "128"),

    ("SERVICOS DE RECARGA DE EXTINTORES", "254"),
    ("SERVIÇOS DE RECARGA DE EXTINTORES", "254"),
    ("RECARGA DE EXTINTORES", "254"),
    ("EXTINTORES", "254"),

    ("MANUTENCAO E PEQUENOS REPAROS", "128"),
    ("MANUTENÇÃO E PEQUENOS REPAROS", "128"),
    ("PEQUENOS REPAROS", "128"),

    ("HIGIENE E LIMPEZA", "94"),
    ("LIMPEZA", "94"),

    ("MOBILIARIOS, EQUIPAMENTOS E UTENSILIOS", "251"),
    ("MOBILIÁRIOS, EQUIPAMENTOS E UTENSÍLIOS", "251"),

    ("MATERIAIS E SERVICOS PEDAGOGICOS", "224"),
    ("MATERIAIS E SERVIÇOS PEDAGÓGICOS", "224"),
    ("MATERIAIS PEDAGOGICOS", "224"),
    ("MATERIAIS PEDAGÓGICOS", "224"),
]
CONTA_RENDIMENTOS = "120"  # RENDIMENTOS DE APLICAÇÕES FINANCEIRAS
CONTA_RECURSOS_PROPRIOS = "202"  # RECURSOS PRÓPRIOS

def limpar_texto(valor):
    if valor is None:
        return ""
    return " ".join(str(valor).replace("\n", " ").split()).strip()


def normalizar(valor):
    valor = limpar_texto(valor).upper()
    troca = {
        "Á": "A", "À": "A", "Â": "A", "Ã": "A",
        "É": "E", "Ê": "E",
        "Í": "I",
        "Ó": "O", "Ô": "O", "Õ": "O",
        "Ú": "U",
        "Ç": "C",
    }
    for a, b in troca.items():
        valor = valor.replace(a, b)
    return valor


def br_para_float(valor):
    if not valor:
        return 0.0
    valor = str(valor).strip().replace(".", "").replace(",", ".")
    try:
        return float(valor)
    except ValueError:
        return 0.0


def float_para_br(valor):
    return float(valor or 0)


def ler_pdf(caminho_pdf):
    reader = PdfReader(caminho_pdf)
    texto = []
    for pagina in reader.pages:
        texto.append(pagina.extract_text() or "")
    return "\n".join(texto)


def extrair_exercicio(texto):
    m = re.search(r"02-Exercício:\s*(\d{4})", texto, re.IGNORECASE)
    if not m:
        m = re.search(r"02-Exercicio:\s*(\d{4})", texto, re.IGNORECASE)
    return m.group(1) if m else str(datetime.now().year)


def extrair_sintese(texto):
    # Usa a primeira ocorrência do BLOCO 2, que contém os totais da receita/despesa.
    bloco = texto
    m = re.search(r"BLOCO 2.*?BLOCO 4", texto, re.IGNORECASE | re.DOTALL)
    if m:
        bloco = m.group(0)

    valores_c = [br_para_float(x) for x in re.findall(r"\(C\)\s*=\s*([\d\.]+,\d{2})", bloco)]
    valores_k = [br_para_float(x) for x in re.findall(r"\(K\)\s*=\s*([\d\.]+,\d{2})", bloco)]

    def soma_indice(i):
        c = valores_c[i] if len(valores_c) > i else 0.0
        k = valores_k[i] if len(valores_k) > i else 0.0
        return c + k

    return {
        "saldo_reprogramado": soma_indice(0),
        "valor_creditado_fnde": soma_indice(1),
        "recursos_proprios": soma_indice(2),
        "rendimento_aplicacao": soma_indice(3),
        "total_receita": soma_indice(4),
    }


def extrair_pagamentos(texto):
    linhas = texto.splitlines()

    pagamentos = []
    atual = []

    marcadores_fim = [
        "PRESTAÇÃO DE CONTAS",
        "DEMONSTRATIVO DA EXECUÇÃO",
        "DIRETORIA FINANCEIRA",
        "COORDENAÇÃO GERAL",
        "BLOCO 1",
        "BLOCO 2",
        "BLOCO 4",
        "TOTAL ",
        "LOCAL E DATA",
        "NOME DO DIRETOR",
    ]

    def finalizar_bloco():
        if not atual:
            return

        bloco_limpo = limpar_texto("\n".join(atual))

        item_match = re.match(r"^(\d{1,3})\s+(.*)$", bloco_limpo)
        if not item_match:
            return

        item = item_match.group(1)
        resto = item_match.group(2)

        datas = re.findall(r"\d{2}/\d{2}/\d{4}", resto)
        valores = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", resto)

        if not datas or not valores:
            return

        data_pagamento = datas[-1]
        valor = br_para_float(valores[-1])

        cnpj_match = re.search(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", resto)
        cnpj = cnpj_match.group(0) if cnpj_match else ""

        nat_match = re.search(
            r"\s([CK])\s+(?:NFEV|NFES|NFS|NFE|NF|GPS|DARF)",
            resto,
            re.IGNORECASE
        )
        nat = nat_match.group(1).upper() if nat_match else ""

        fornecedor = ""
        material = ""

        if cnpj:
            antes, depois = resto.split(cnpj, 1)
            fornecedor = limpar_texto(antes)

            if nat:
                partes = re.split(
                    rf"\s{nat}\s+(?:NFEV|NFES|NFS|NFE|NF|GPS|DARF)",
                    depois,
                    flags=re.IGNORECASE
                )
                material = limpar_texto(partes[0])
            else:
                material = limpar_texto(depois)
        else:
            material = re.sub(
                r"\d{1,3}(?:\.\d{3})*,\d{2}.*",
                "",
                resto
            )
            material = limpar_texto(material)

        pagamentos.append({
            "item": item,
            "fornecedor": fornecedor,
            "cnpj": cnpj,
            "material": material,
            "natureza": nat,
            "data_pagamento": data_pagamento,
            "valor": valor,
        })

    for linha in linhas:
        linha = linha.strip()

        if not linha:
            continue

        linha_norm = normalizar(linha)

        if any(m in linha_norm for m in marcadores_fim):
            finalizar_bloco()
            atual = []
            continue

        if re.match(r"^\d{1,3}\s+", linha):
            finalizar_bloco()
            atual = [linha]
        else:
            if atual:
                atual.append(linha)

    finalizar_bloco()

    unicos = []
    chaves = set()

    for p in pagamentos:
        chave = (p["item"], p["data_pagamento"], p["valor"])

        if chave not in chaves:
            chaves.add(chave)
            unicos.append(p)

    return unicos


def extrair_plano_contas(caminho_plano_pdf):
    texto = ler_pdf(caminho_plano_pdf)
    contas = {}

    for linha in texto.splitlines():
        linha = limpar_texto(linha)
        m = re.match(r"^(\d+)\s+([1-9]\d*(?:\.\d+)+)\s+(.+)$", linha)
        if not m:
            continue
        reduzida, conta, descricao = m.groups()
        contas[reduzida] = {
            "reduzida": reduzida,
            "conta": conta,
            "descricao": descricao,
            "descricao_norm": normalizar(descricao),
        }

    return contas

def grupo_da_conta(conta_reduzida, plano_contas):
    conta_reduzida = str(conta_reduzida).strip()

    conta = plano_contas.get(conta_reduzida)

    if conta:
        numero_contabil = conta.get("conta", "")

        if numero_contabil.startswith("3."):
            return "despesa"

        if numero_contabil.startswith("4."):
            return "receita"

    # Correção manual para contas que o PDF pode não ler direito
    RECEITAS_PDDE = {
        "125", "126", "151", "152", "154", "159", "161",
        "187", "194", "195", "196", "198", "201", "204",
        "206", "208", "209", "217", "219", "221", "223",
        "227", "229", "231", "233", "235", "237", "239",
        "242", "243", "245", "247", "249", "252",
        "116", "117", "120", "202", "212"
    }

    DESPESAS_PDDE = {
        "97", "98", "251", "82", "83", "84", "85", "86",
        "87", "127", "128", "160", "165", "166", "254",
        "65", "66", "67", "68", "73", "74", "75", "76",
        "177", "69", "70", "71", "72", "183", "184",
        "135", "80", "157", "158", "77", "78", "162",
        "163", "81", "88", "89", "90", "91", "92", "93",
        "94", "95", "96", "136", "101", "102", "190",
        "42", "53", "52", "180", "182", "188", "189",
        "253", "106", "107", "153", "175", "211"
    }

    if conta_reduzida in RECEITAS_PDDE:
        return "receita"

    if conta_reduzida in DESPESAS_PDDE:
        return "despesa"

    return ""

def validar_conta_por_tipo(conta_reduzida, tipo, plano_contas):
    grupo = grupo_da_conta(conta_reduzida, plano_contas)

    if tipo == "despesa" and grupo != "despesa":
        return False

    if tipo == "receita" and grupo != "receita":
        return False

    return True


def garantir_conta_valida(conta_reduzida, tipo, plano_contas, fallback):
    if validar_conta_por_tipo(conta_reduzida, tipo, plano_contas):
        return conta_reduzida

    if validar_conta_por_tipo(fallback, tipo, plano_contas):
        return fallback

    raise ValueError(
        f"Conta inválida para {tipo}: {conta_reduzida}. "
        f"Também não foi possível usar a conta padrão {fallback}."
    )

def classificar_conta_despesa(pagamento, plano_contas):
    texto = normalizar(
        f"{pagamento.get('fornecedor','')} {pagamento.get('material','')}"
    )

    conta_encontrada = ""

    for palavra, conta in REGRAS_CONTAS_DESPESAS:
        if normalizar(palavra) in texto:
            conta_encontrada = conta
            break

    if not conta_encontrada:
        if pagamento.get("natureza") == "K":
            conta_encontrada = "15"
        else:
            conta_encontrada = "180"

    return garantir_conta_valida(
        conta_encontrada,
        "despesa",
        plano_contas,
        "180"
    )

def classificar_conta_receita(lancamento, plano_contas, conta_receita_pdde):
    texto = normalizar(
        f"{lancamento.get('historico','')} {lancamento.get('descricao','')}"
    )

    if (
        "RENDIMENTO" in texto
        or "APLICACAO" in texto
        or "APLICAÇÃO" in texto
        or "INVESTIMENTO" in texto
    ):
        return garantir_conta_valida(
            CONTA_RENDIMENTOS,
            "receita",
            plano_contas,
            CONTA_RENDIMENTOS
        )

    if (
        "RECURSOS PROPRIOS" in texto
        or "RECURSOS PRÓPRIOS" in texto
        or "RECURSO PROPRIO" in texto
        or "RECURSO PRÓPRIO" in texto
    ):
        return garantir_conta_valida(
            CONTA_RECURSOS_PROPRIOS,
            "receita",
            plano_contas,
            CONTA_RECURSOS_PROPRIOS
        )

    return garantir_conta_valida(
        conta_receita_pdde,
        "receita",
        plano_contas,
        conta_receita_pdde
    )
def historico_pagamento(pagamento):
    if pagamento.get("fornecedor"):
        return pagamento["fornecedor"]
    if pagamento.get("material"):
        return pagamento["material"]
    return "DESPESA PDDE"


def escrever_linha(sheet, linha, data, debito, credito, valor, historico, saldo):
    sheet.cell(row=linha, column=1).value = data
    sheet.cell(row=linha, column=2).value = str(debito) if debito else ""
    sheet.cell(row=linha, column=3).value = str(credito) if credito else ""
    sheet.cell(row=linha, column=4).value = float_para_br(valor) if valor not in (None, "") else ""
    sheet.cell(row=linha, column=5).value = historico
    sheet.cell(row=linha, column=6).value = float_para_br(saldo)


def gerar_planilha_importacao(
    caminho_pdde_pdf,
    caminho_plano_pdf,
    caminho_modelo_xlsx,
    caminho_saida_xlsx,
    conta_caixa,
    conta_receita_pdde,
):
    texto_pdde = ler_pdf(caminho_pdde_pdf)
    exercicio = extrair_exercicio(texto_pdde)
    sintese = extrair_sintese(texto_pdde)
    pagamentos = extrair_pagamentos(texto_pdde)
    plano = extrair_plano_contas(caminho_plano_pdf)

    conta_receita_pdde = str(conta_receita_pdde).strip()

    if not validar_conta_por_tipo(conta_receita_pdde, "receita", plano):
        return {"400": f"A conta de receita informada ({conta_receita_pdde}) não pertence ao grupo 4 - Receitas."}

    CONTA_RENDIMENTOS_VALIDADA = garantir_conta_valida(
        CONTA_RENDIMENTOS,
        "receita",
        plano,
        "120"
    )

    CONTA_RECURSOS_PROPRIOS_VALIDADA = garantir_conta_valida(
        CONTA_RECURSOS_PROPRIOS,
        "receita",
        plano,
        "202"
    )

    if caminho_modelo_xlsx and os.path.exists(caminho_modelo_xlsx):
        wb = load_workbook(caminho_modelo_xlsx)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "JANEIRO"
        ws.append(["Data", "Debito", "Crédito", "Valor", "Histórico", "Saldo"])

    # Limpa linhas antigas abaixo do cabeçalho
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    saldo = sintese["saldo_reprogramado"]
    linha = 2

    # Saldo inicial
    data_saldo = f"01/01/{exercicio}"
    escrever_linha(ws, linha, data_saldo, "", "", "", "SALDO INICIAL", saldo)
    linha += 1

    data_receita = f"01/01/{exercicio}"

    if sintese["valor_creditado_fnde"] > 0:
        saldo += sintese["valor_creditado_fnde"]
        escrever_linha(
            ws, linha, data_receita,
            conta_caixa, conta_receita_pdde,
            sintese["valor_creditado_fnde"],
            "REPASSE PDDE - FNDE",
            saldo
        )
        linha += 1

    if sintese["recursos_proprios"] > 0:
        saldo += sintese["recursos_proprios"]
        escrever_linha(
            ws, linha, data_receita,
            conta_caixa, CONTA_RECURSOS_PROPRIOS_VALIDADA,
            sintese["recursos_proprios"],
            "RECURSOS PRÓPRIOS PDDE",
            saldo
        )
        linha += 1

    if sintese["rendimento_aplicacao"] > 0:
        saldo += sintese["rendimento_aplicacao"]
        escrever_linha(
            ws, linha, data_receita,
            conta_caixa, CONTA_RENDIMENTOS_VALIDADA,
            sintese["rendimento_aplicacao"],
            "RENDIMENTOS DE APLICAÇÕES FINANCEIRAS",
            saldo
        )
        linha += 1

    # Despesas
    avisos = []

    for p in pagamentos:
        conta_despesa = classificar_conta_despesa(p, plano)
        saldo -= p["valor"]

        if conta_despesa not in plano:
            avisos.append(f"Conta {conta_despesa} não encontrada no plano para item {p['item']}")

        escrever_linha(
            ws, linha,
            p["data_pagamento"],
            conta_despesa,
            conta_caixa,
            p["valor"],
            historico_pagamento(p),
            saldo
        )
        linha += 1

    # Formatação básica
    for col, width in {
        "A": 14,
        "B": 12,
        "C": 12,
        "D": 14,
        "E": 45,
        "F": 14,
    }.items():
        ws.column_dimensions[col].width = width

    for r in range(2, linha):
        ws.cell(r, 4).number_format = '#,##0.00'
        ws.cell(r, 6).number_format = '#,##0.00'

    wb.save(caminho_saida_xlsx)

    return {
        "arquivo": caminho_saida_xlsx,
        "pagamentos": len(pagamentos),
        "saldo_inicial": sintese["saldo_reprogramado"],
        "receita_fnde": sintese["valor_creditado_fnde"],
        "rendimento": sintese["rendimento_aplicacao"],
        "avisos": avisos,
    }
    
def gerar_planilha_manual_pdde(
    caminho_plano_pdf,
    caminho_modelo_xlsx,
    caminho_saida_xlsx,
    conta_caixa,
    conta_receita_pdde,
    saldo_inicial,
    lancamentos
):
    plano = extrair_plano_contas(caminho_plano_pdf)

    conta_receita_pdde = str(conta_receita_pdde).strip()

    if not validar_conta_por_tipo(conta_receita_pdde, "receita", plano):
        return {"400": f"A conta de receita informada ({conta_receita_pdde}) não pertence ao grupo 4 - Receitas."}

    if caminho_modelo_xlsx and os.path.exists(caminho_modelo_xlsx):
        wb = load_workbook(caminho_modelo_xlsx)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "JANEIRO"
        ws.append(["Data", "Debito", "Crédito", "Valor", "Histórico", "Saldo"])

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    saldo = saldo_inicial
    linha = 2

    escrever_linha(
        ws,
        linha,
        "",
        "",
        "",
        "",
        "SALDO INICIAL",
        saldo
    )

    linha += 1

    receitas = [l for l in lancamentos if l["tipo"] == "receita"]
    despesas = [l for l in lancamentos if l["tipo"] == "despesa"]

    for l in receitas:
        saldo += l["valor"]

        conta_receita = classificar_conta_receita(
            l,
            plano,
            conta_receita_pdde
        )

        data = datetime.strptime(l["data"], "%Y-%m-%d").strftime("%d/%m/%Y")

        escrever_linha(
            ws,
            linha,
            data,
            conta_caixa,
            conta_receita,
            l["valor"],
            l["historico"],
            saldo
        )

        linha += 1

    for l in despesas:
        saldo -= l["valor"]

        pagamento_fake = {
            "fornecedor": l["historico"],
            "material": l["descricao"],
            "natureza": ""
        }

        conta_despesa = classificar_conta_despesa(
            pagamento_fake,
            plano
        )

        data = datetime.strptime(l["data"], "%Y-%m-%d").strftime("%d/%m/%Y")

        escrever_linha(
            ws,
            linha,
            data,
            conta_despesa,
            conta_caixa,
            l["valor"],
            l["historico"],
            saldo
        )

        linha += 1

    for col, width in {
        "A": 14,
        "B": 12,
        "C": 12,
        "D": 14,
        "E": 45,
        "F": 14,
    }.items():
        ws.column_dimensions[col].width = width

    for r in range(2, linha):
        ws.cell(r, 4).number_format = '#,##0.00'
        ws.cell(r, 6).number_format = '#,##0.00'

    wb.save(caminho_saida_xlsx)

    return caminho_saida_xlsx
